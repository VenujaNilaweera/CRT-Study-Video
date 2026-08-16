from __future__ import annotations

import json
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "data.xlsx"
SHEET_NAME = "Annotations"
# 0.0.0.0 = accept connections from other devices on the same WiFi (e.g. a
# phone opening the page from the PC's LAN IP). Use "127.0.0.1" to restrict
# to this machine only.
HOST = "0.0.0.0"
PORT = 8787

HEADERS = [
    "Name",
    "Role",
    "Age Group",
    "Collection",
    "Video #",
    "Video Title",
    "Time (s)",
    "FPS",
    "Frame #",
    "Submitted At",
]


def ensure_sheet():
    if WORKBOOK_PATH.exists():
        workbook = load_workbook(WORKBOOK_PATH)
    else:
        workbook = Workbook()

    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        sheet = workbook.active
        sheet.title = SHEET_NAME

    if sheet.max_row == 1 and all(sheet.cell(1, col).value is None for col in range(1, len(HEADERS) + 1)):
        sheet.append(HEADERS)
    elif [sheet.cell(1, col).value for col in range(1, len(HEADERS) + 1)] != HEADERS:
        sheet.insert_rows(1)
        for col, header in enumerate(HEADERS, start=1):
            sheet.cell(1, col).value = header

    header_fill = PatternFill("solid", fgColor="0E2A2F")
    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = [22, 16, 12, 18, 10, 24, 12, 8, 10, 22][col - 1]

    # Self-heal: drop any leftover duplicate header rows below row 1
    # (e.g. left behind when the header layout changed between versions).
    for row_idx in range(sheet.max_row, 1, -1):
        values = [sheet.cell(row_idx, col).value for col in range(1, len(HEADERS) + 1)]
        if values == HEADERS:
            sheet.delete_rows(row_idx, 1)

    sheet.freeze_panes = "A2"
    return workbook, sheet


def append_annotation(payload: dict) -> int:
    workbook, sheet = ensure_sheet()
    submitted_at = payload.get("submittedAt") or datetime.now().isoformat(timespec="seconds")
    row = [
        payload.get("name", ""),
        payload.get("status", ""),
        payload.get("age", ""),
        payload.get("collection", ""),
        payload.get("videoIndex", ""),
        payload.get("videoTitle", ""),
        payload.get("timeSeconds", ""),
        payload.get("fps", ""),
        payload.get("frameIndex", ""),
        submitted_at,
    ]
    sheet.append(row)
    sheet.cell(sheet.max_row, 7).number_format = "0.000"
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(WORKBOOK_PATH)
    return sheet.max_row


class Handler(BaseHTTPRequestHandler):
    def _send_json(self, status: int, body: dict):
        data = json.dumps(body).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "content-type")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.end_headers()
        self.wfile.write(data)

    def do_OPTIONS(self):
        self._send_json(200, {"ok": True})

    def do_GET(self):
        if self.path == "/health":
            self._send_json(200, {"ok": True, "workbook": str(WORKBOOK_PATH)})
        else:
            self._send_json(404, {"ok": False, "error": "not found"})

    def do_POST(self):
        if self.path != "/annotation":
            self._send_json(404, {"ok": False, "error": "not found"})
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            row_number = append_annotation(payload)
            self._send_json(200, {"ok": True, "row": row_number, "workbook": str(WORKBOOK_PATH)})
        except Exception as exc:
            self._send_json(500, {"ok": False, "error": str(exc)})

    def log_message(self, format, *args):
        return


def _lan_ip() -> str:
    """Best-effort LAN IP of this PC (the address a phone should use)."""
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))  # no packets sent; just picks the outbound iface
        return s.getsockname()[0]
    except Exception:
        return "127.0.0.1"
    finally:
        s.close()


if __name__ == "__main__":
    ensure_sheet()[0].save(WORKBOOK_PATH)
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    ip = _lan_ip()
    print(f"Excel helper listening on {HOST}:{PORT}")
    print(f"  this PC:  http://127.0.0.1:{PORT}")
    print(f"  phone / other devices on the same WiFi:  http://{ip}:{PORT}")
    print(f"Updating workbook: {WORKBOOK_PATH}")
    server.serve_forever()
